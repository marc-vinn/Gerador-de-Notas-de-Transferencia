import io
import os
import json
import pytest
import openpyxl
from api.index import app
from core.tests.test_parser import FIXTURE_FILE_PATH

@pytest.fixture
def client():
    app.config["TESTING"] = True
    with app.test_client() as client:
        yield client

def test_api_index_route(client):
    rv = client.get("/")
    assert rv.status_code in [200, 404]

def test_api_upload_endpoint(client):
    with open(FIXTURE_FILE_PATH, "rb") as f:
        file_bytes = f.read()

    data = {
        "file": (io.BytesIO(file_bytes), "relatorio.xls")
    }
    rv = client.post("/api/upload", data=data, content_type="multipart/form-data")
    assert rv.status_code == 200
    res_json = rv.get_json()
    assert res_json["success"] is True
    assert res_json["summary"]["item_count"] == 72
    assert len(res_json["products"]) == 72

def test_api_analyze_multi_endpoint_with_custom_stock_schema(client):
    """
    Tests /api/analyze-multi with stock reports formatted as:
    Col A: SKU, Col B: Descrição, Col C: Estoque Atual, Col D: Unidades Reservadas
    """
    # Create Sales Workbook
    wb_sales = openpyxl.Workbook()
    ws_s = wb_sales.active
    ws_s.append(["Código", "Descrição", "Quantidade", "Valor Líquido"])
    ws_s.append(["SKU-001", "Camisa Polo Azul", 10, 500.0]) # 10/sem -> Need 40
    stream_sales = io.BytesIO()
    wb_sales.save(stream_sales)
    sales_bytes = stream_sales.getvalue()

    # Create Branch Stock Workbook (Col A, B, C, D)
    wb_b_stock = openpyxl.Workbook()
    ws_bs = wb_b_stock.active
    ws_bs.append(["SKU", "Descrição", "Estoque Atual", "Unidades Reservadas"])
    ws_bs.append(["SKU-001", "Camisa Polo Azul", 15, 5]) # 15 - 5 = 10 available. Deficit = 30.
    stream_b_stock = io.BytesIO()
    wb_b_stock.save(stream_b_stock)
    b_stock_bytes = stream_b_stock.getvalue()

    # Create Matrix Sales Workbook
    wb_m_sales = openpyxl.Workbook()
    ws_ms = wb_m_sales.active
    ws_ms.append(["Código", "Descrição", "Quantidade", "Valor Líquido"])
    ws_ms.append(["SKU-001", "Camisa Polo Azul", 20, 1000.0]) # 30d demand = 20
    stream_m_sales = io.BytesIO()
    wb_m_sales.save(stream_m_sales)
    m_sales_bytes = stream_m_sales.getvalue()

    # Create Matrix Stock Workbook (Col A, B, C, D)
    wb_m_stock = openpyxl.Workbook()
    ws_ms_stock = wb_m_stock.active
    ws_ms_stock.append(["SKU", "Descrição", "Estoque Atual", "Unidades Reservadas"])
    ws_ms_stock.append(["SKU-001", "Camisa Polo Azul", 70, 10]) # 70 - 10 = 60 available. Balance = 60 - 20 = 40 >= 30.
    stream_m_stock = io.BytesIO()
    wb_m_stock.save(stream_m_stock)
    m_stock_bytes = stream_m_stock.getvalue()

    data = {
        "branch_sales": (io.BytesIO(sales_bytes), "filial_vendas.xlsx"),
        "branch_stock": (io.BytesIO(b_stock_bytes), "filial_estoque.xlsx"),
        "matrix_sales": (io.BytesIO(m_sales_bytes), "matriz_vendas.xlsx"),
        "matrix_stock": (io.BytesIO(m_stock_bytes), "matriz_estoque.xlsx")
    }
    rv = client.post("/api/analyze-multi", data=data, content_type="multipart/form-data")
    assert rv.status_code == 200
    res_json = rv.get_json()
    assert res_json["success"] is True
    assert len(res_json["approved_normal"]) == 1
    item = res_json["approved_normal"][0]
    assert item["sku"] == "SKU-001"
    assert item["quantity"] == 30.0
    assert item["branch_stock"] == 10.0
    assert item["matrix_stock"] == 60.0


def test_api_analyze_multi_with_real_sample_files(client):
    """
    Tests /api/analyze-multi with the real sample files from Teste_com_exemplo_real.
    """
    folder = os.path.join(os.path.dirname(__file__), "..", "..", "Teste_com_exemplo_real")
    f_b_sales = os.path.join(folder, "Relatorio de vendas dos ultimos 7 dias Derya.xls")
    f_b_stock = os.path.join(folder, "Saldo estoque atual Derya.xls")
    f_m_sales = os.path.join(folder, "Relatorio de vendas Arboretho últimos 60 dias.xls")
    f_m_stock = os.path.join(folder, "Saldo estoque atual Arboretho.xls")

    if not os.path.exists(f_b_sales):
        pytest.skip("Arquivos reais de teste não encontrados na pasta local.")

    with open(f_b_sales, "rb") as f1, open(f_b_stock, "rb") as f2, open(f_m_sales, "rb") as f3, open(f_m_stock, "rb") as f4:
        data = {
            "branch_sales": (io.BytesIO(f1.read()), "vendas_derya.xls"),
            "branch_stock": (io.BytesIO(f2.read()), "estoque_derya.xls"),
            "matrix_sales": (io.BytesIO(f3.read()), "vendas_arboretho.xls"),
            "matrix_stock": (io.BytesIO(f4.read()), "estoque_arboretho.xls")
        }
        rv = client.post("/api/analyze-multi", data=data, content_type="multipart/form-data")
        assert rv.status_code == 200
        res_json = rv.get_json()
        assert res_json["success"] is True
        assert res_json["summary"]["normal_items_count"] == 8
        assert res_json["summary"]["purchase_alerts_count"] == 28
        assert res_json["summary"]["reverse_items_count"] == 11


def test_api_generate_xml_security_gate_blocks_unconfigured_companies(client):
    """Verifies that generate-xml strictly rejects requests if companies are not configured."""
    with open(FIXTURE_FILE_PATH, "rb") as f:
        file_bytes = f.read()

    # Case 1: No companies provided at all
    data = {"file": (io.BytesIO(file_bytes), "relatorio.xls")}
    rv = client.post("/api/generate-xml", data=data, content_type="multipart/form-data")
    assert rv.status_code == 400
    res_json = rv.get_json()
    assert res_json["success"] is False
    assert "Empresa Emitente (Matriz) não configurada" in res_json["error"]

    # Case 2: Emitter provided, but recipient missing
    data = {
        "file": (io.BytesIO(file_bytes), "relatorio.xls"),
        "emitter_cnpj": "11.222.333/0001-44",
        "emitter_name": "MATRIZ ARBORETHO LTDA",
        "emitter_uf": "GO"
    }
    rv = client.post("/api/generate-xml", data=data, content_type="multipart/form-data")
    assert rv.status_code == 400
    res_json = rv.get_json()
    assert res_json["success"] is False
    assert "Empresa Destinatária (Filial) não configurada" in res_json["error"]

def test_api_generate_xml_normal_direction_with_dual_companies(client):
    """Tests normal transfer (Matrix -> Branch) with both companies registered."""
    with open(FIXTURE_FILE_PATH, "rb") as f:
        file_bytes = f.read()

    data = {
        "file": (io.BytesIO(file_bytes), "relatorio.xls"),
        "direction": "matrix_to_branch",
        "emitter_cnpj": "11.222.333/0001-44",
        "emitter_name": "MATRIZ ARBORETHO LTDA",
        "emitter_uf": "GO",
        "recipient_cnpj": "99.888.777/0001-11",
        "recipient_name": "FILIAL GOIANIA LTDA",
        "recipient_uf": "GO"
    }
    rv = client.post("/api/generate-xml", data=data, content_type="multipart/form-data")
    assert rv.status_code == 200
    assert rv.mimetype == "application/xml"
    xml_data = rv.data.decode("utf-8")
    assert "<nfeProc" in xml_data
    assert "nfe_transferencia_" in rv.headers.get("Content-Disposition", "")
    assert "MATRIZ ARBORETHO LTDA" in xml_data
    assert "FILIAL GOIANIA LTDA" in xml_data

def test_api_generate_xml_reverse_direction(client):
    """Tests generating reverse transfer XML (Branch -> Matrix) with symmetric swap."""
    edited_products = [
        {
            "sku": "REV-SKU-001",
            "description": "PRODUTO EXCEDENTE FILIAL",
            "quantity": 15.0,
            "unit_price": 50.0,
            "total_price": 750.0
        }
    ]
    data = {
        "products": json.dumps(edited_products),
        "filename": "excedente_filial.xls",
        "direction": "branch_to_matrix",
        "emitter_cnpj": "11.222.333/0001-44",
        "emitter_name": "MATRIZ ARBORETHO LTDA",
        "emitter_uf": "GO",
        "recipient_cnpj": "99.888.777/0001-11",
        "recipient_name": "FILIAL GOIANIA LTDA",
        "recipient_uf": "GO"
    }
    rv = client.post("/api/generate-xml", data=data, content_type="multipart/form-data")
    assert rv.status_code == 200
    assert rv.mimetype == "application/xml"
    xml_data = rv.data.decode("utf-8")
    assert "nfe_transferencia_reversa" in rv.headers.get("Content-Disposition", "")
    # In reverse direction: Emitter is the Branch (Filial), Recipient is the Matrix (Matriz)
    assert "<emit>" in xml_data and "FILIAL GOIANIA LTDA" in xml_data
    assert "<dest>" in xml_data and "MATRIZ ARBORETHO LTDA" in xml_data
    assert "<qCom>15.0000</qCom>" in xml_data
    assert "<vUnCom>50.00</vUnCom>" in xml_data
    assert "<vProd>750.00</vProd>" in xml_data


